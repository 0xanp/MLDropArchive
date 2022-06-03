import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import pathlib
import os


# ---- LOAD ALL DATA PATHS ----
directory = os.fsencode("data")
paths= {}
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    if filename.endswith(".xlsx"):
        paths[filename.replace('.xlsx','')] = "data/"+filename
        continue
    else:
        continue

# ---- HYPERLINK ----
def make_clickable(link):
    # target _blank to open new window
    # extract clickable text to display for your link
    text = link.split('=')[0]
    return f'<a target="_blank" href="{link}">{text}</a>'

# ---- TABLE STYLING ----
def coloring(s):
    color = 'green' if  s == 'green' else 'yellow' if s == 'yellow' else 'orange'
    return f'color: {color}'

# ---- READ EXCEL ----
@st.cache
def load_data(path,calendar_tag):
    excel_file = path 
    wb = load_workbook(excel_file, data_only = True)
    sh = wb[wb.sheetnames[0]]
    df = pd.read_excel(excel_file,
                    sheet_name=wb.sheetnames[0],
                    usecols='A:D',skiprows=[1])
    
    df = df[df['Comments'].notnull()]
    color = []
    cycle = []
    for i in df.index:
        hex = sh['A'+str(i+3)].fill.start_color.index # this gives you Hexadecimal value of the color
        cycle.append(calendar_tag)
        if hex == 'FF00FF00':
            color.append('green')
        elif hex == 'FFFFFF00':
            color.append('yellow')
        elif hex == 'FFFF9900':
            color.append('orange')
    df['Twitter'] = df['Twitter'].apply(make_clickable)
    df['Color'] = color
    df['Cycle'] = cycle
    df['Comments'] = df['Comments'].str.replace("\n\n","").str.replace("\n \n","")
    df['Mint Date'] = df['Mint Date'].astype(str)
    return df

# ---- PAGE CONFIG ----
st.set_page_config(page_title='ML Drop Calendar Archive',page_icon=":waxing_crescent_moon:", layout="wide")

# ---- HEADER ----
st.write(":waxing_crescent_moon: :first_quarter_moon: :waxing_gibbous_moon: :waxing_gibbous_moon: :full_moon: :waning_gibbous_moon: :last_quarter_moon: :waning_crescent_moon:")
st.title('Midnightlabs Drop Calendar Archive')
st.write(":waxing_crescent_moon: :first_quarter_moon: :waxing_gibbous_moon: :waxing_gibbous_moon: :full_moon: :waning_gibbous_moon: :last_quarter_moon: :waning_crescent_moon:")
st.markdown("""
    <p>
    <strong> Description: </strong> Projects highlighted in
    <span style="color: #32CD32">green</span>
    were found to have strong fundumentals and potential to be mid-to-long term holds. Projects in
    <span style="color: #FFFF00">yellow</span> 
    are on our watchlist and will continue to be monitored as they develop. Projects highlighted in
    <span style="color: #FF4500">orange</span> 
    show signs of potential but lack important information needed to make a final call.
    </p>
""",unsafe_allow_html=True)


# ---- LOAD ALL DATA ----
df = pd.DataFrame()
for name in paths.keys():
    df1 = load_data(paths[name],name)
    df = pd.concat([df, df1], ignore_index=True)
    df.index = np.arange(1, len(df)+1)

# ---- SIDE BAR ----
color = st.sidebar.multiselect(
    "Filter the color:",
    options=df["Color"].unique(),
    default=df["Color"].unique()
)
cycle = st.sidebar.multiselect(
    "Select the cycle:",
    options=df["Cycle"].unique(),
    default=df["Cycle"].unique()
)

df_selection = df.query(
    "Color == @color & Cycle == @cycle"
)

# ---- Hiding default wartermark ----
hide_menu_style = """
        <style>
        #MainMenu {visibility: hidden; }
        footer {visibility: hidden;}
        </style>
        """
st.markdown(hide_menu_style, unsafe_allow_html=True)

st.write(df_selection[['Color','Project','Twitter','Comments','Mint Date']].sort_values(by=['Color']).style.applymap(coloring, subset=['Color']).hide().to_html(escape=False, index=False), unsafe_allow_html=True)